# -*- coding: UTF-8 -*-
"""
@File: read_excel.py
@Description: 
@Author: lee sure
@DateTime: 2021/11/08 02:42
@Project_name: pythonProject
@IDE: PyCharm
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from openpyxl.styles.fills import FILL_SOLID




file= u"op_excel.xlsx"
workbook = load_workbook(file)

worksheet = workbook[workbook.sheetnames[0]]

row_list = [[it.value for it in item] for item in worksheet.rows]
print(row_list)

col_list = [[j.value for j in items] for items in worksheet.columns]
print(col_list)

data = worksheet.cell(row=1, column=1)
print(data.value)

if "new" not in workbook.sheetnames:
    worksheet = workbook.create_sheet(title="new", index=0)
else:
    worksheet = workbook[workbook.sheetnames[0]]

worksheet['A1'].value = 1231
font = Font(name='Calibri',
            size=12,
            bold=True,
            italic=True,
            vertAlign=None,
            underline=Font.UNDERLINE_SINGLE,
            strike=False,
            color=colors.COLOR_INDEX[3])

fill = PatternFill(fill_type=FILL_SOLID,
                   fgColor='00FF0000')
worksheet['A1'].font = font
worksheet['A1'].fill = fill

workbook.save(file)
workbook.close()




data= pd.read_excel("op_excel.xlsx",sheet_name=1)
# 读取后drop
mydata = data.drop([1,3], axis=0)
# 保存新的数据
book = load_workbook('data.xlsx')
writer = pd.ExcelWriter('data.xlsx',engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
mydata.to_excel(writer, "mysheet")
writer.save()
